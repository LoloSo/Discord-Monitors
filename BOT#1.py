from telegram import Update
from telegram.ext import Updater, MessageHandler, Filters, CommandHandler, CallbackContext
from openpyxl import load_workbook

wb = load_workbook('database.xlsx')
vendor = wb['articels']

token = '1684008354:AAGtETeI2tgXd-AM_6AQvJmh43aTSYjkXxE'


def main():

    updater = Updater(token=token)

    dispatcher = updater.dispatcher

    handler = MessageHandler(Filters.all, do_echo)
    start_handler = CommandHandler('start', do_start)
    help_handler = CommandHandler('help', do_help)
    #add_vendor_code_handler = CommandHandler('addven', do_add_vendor)
    #delete_vendor_code_handler = CommandHandler('delven', do_delete_vendor)

    dispatcher.add_handler(start_handler)
    dispatcher.add_handler(help_handler)
    #dispatcher.add_handler(add_vendor_code_handler)
    #dispatcher.add_handler(delete_vendor_code_handler)
    dispatcher.add_handler(handler)

    updater.start_polling()
    updater.idle()


def do_echo(update, context: CallbackContext):
    if context.user_data:
        command = context.user_data['команда']
        if command == '/addven':
            do_add_vendor(update, context)
            return command
        elif command == '/delven':
            do_delete_vendor(update, context)
            return command

    update.message.reply_text(text='Done!')


def do_start(update, context):

    update.message.reply_text(
        text='Hi',
    )


def do_help(update: Update, context: CallbackContext):
    user_id = update.message.from_user.id
    name = update.message.from_user.first_name
    update.message.reply_text(text=f'Привет, {name}!\nТвой user_id: {user_id}\nПомощь на подходе')


def do_words(update: Update, context: CallbackContext):

    context.user_data['команда'] = do_echo(update, context)

    if context.user_data['команда'] == '/addven':# or update.message.text == '/delven':
        update.message.reply_text(text='Введите артикул: ')
        return update.message.text
    elif context.user_data['команда'] == '/delven':
        update.message.reply_text(text='Для удаленияЖ ')
        return update.message.text
    else:
        update.message.reply_text(text='sry...')


def do_add_vendor(update: Update, context: CallbackContext):

    words = do_words(update, context)

    for i in range(2, 100):

        if vendor.cell(column=1, row=i).value is None:
            vendor.cell(column=1, row=i).value = words
            wb.save('database.xlsx')
            break

    return wb.save('database.xlsx')


def do_delete_vendor(update: Update, context: CallbackContext):

    update.message.reply_text(text='Введите артикул для удаления: ')

    for i in range(2, 100):

        if vendor.cell(column=1, row=i).value == do_words(update, context):
            vendor.cell(column=1, row=i).value = None
            wb.save('database.xlsx')
            break

    return wb.save('database.xlsx')

main()