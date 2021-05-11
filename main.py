import openpyxl as p

wb = p.load_workbook('database.xlsx')
articles = wb['articels']
command = input('cmd: ')

if command == 'start':
    art = input('Ведите артикул: ')
    for i in range(1, 100):
        if articles.cell(row=i, column=1).value is None:
            articles.cell(row=i, column=1).value = art
            wb.save('database.xlsx')
            break
