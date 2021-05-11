from openpyxl import load_workbook

wb = load_workbook('database.xlsx')
vendor = wb['articels']

def del_vendor():

    vendor_text = input('Введите артикул: ')

    for i in range(1, 100):
        if vendor.cell(column=1, row=i).value == vendor_text:
            vendor.cell(column=1, row=i).value = None
            wb.save('database.xlsx')
            break
del_vendor()